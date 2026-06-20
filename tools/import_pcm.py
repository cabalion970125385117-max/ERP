#!/usr/bin/env python3
"""
ATCA-ERP — MOD-28 PCM importer.

Reads ATC-PCM-001 Rev A.xlsx (Process Capability Matrix) and emits:
  - tools/pcm_demo.json  : demo payloads for /mod28/* (merged into atca-demo.js
                            and loaded by preview_server.py)
Run:  python tools/import_pcm.py "<path to ATC-PCM-001 REV A.xlsx>"

The MasterList sheet is a merged-cell matrix:
  A=Process  B=Customer/Category  C=Specification  D=Tier1 F=Tier2 H=Tier3 J=Tier4
  L=MaxLen  M=MaxWid  N=MaxDep   (E,G,I,K are merge-spill blanks)
Process and Customer cells are forward-filled; each row carrying a Specification
becomes one capability record.
"""
import sys, json, re, os

DEFAULT_XLSX = os.path.expanduser(r'~\Downloads\ATC-PCM-001 REV A.xlsx')

GROUP_MAP = {  # process_name (lowercased contains) -> group
    'anodiz': 'ELECTROPLATING', 'black oxide': 'ELECTROPLATING', 'chromate': 'ELECTROPLATING',
    'plating': 'ELECTROPLATING', 'electroless': 'ELECTROPLATING', 'nickel': 'ELECTROPLATING',
    'phosphat': 'ELECTROPLATING', 'passivation': 'ELECTROPLATING', 'chromium': 'ELECTROPLATING',
    'cadmium': 'ELECTROPLATING', 'electropolish': 'ELECTROPLATING', 'blueing': 'ELECTROPLATING',
    'chem clean': 'ELECTROPLATING', 'ndt': 'NDT', 'penetrant': 'NDT', 'magnetic': 'NDT',
    'cleanroom': 'CLEANROOM', 'coating': 'COATING',
}

def group_for(name):
    n = (name or '').lower()
    for k, g in GROUP_MAP.items():
        if k in n:
            return g
    return 'OTHER'

def clean(v):
    if v is None:
        return None
    s = str(v).replace('\n', ' ').strip()
    return s or None

def first_line(v):
    if v is None:
        return None
    return str(v).split('\n')[0].strip()[:120] or None

def main():
    import openpyxl
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['MasterList']

    # find header row (the one containing 'Process' in col A)
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        if row and isinstance(row[0], str) and row[0].strip().lower() == 'process':
            header_row = i
            break
    header_row = header_row or 5

    processes = {}   # name -> dict
    caps = []
    cur_proc = cur_cust = None
    next_cap = 1
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        a = clean(row[0]) if len(row) > 0 else None      # Process
        b = clean(row[1]) if len(row) > 1 else None      # Customer/category
        c = clean(row[2]) if len(row) > 2 else None       # Specification
        if a:
            cur_proc = first_line(a)
            cur_cust = None
        if b:
            cur_cust = first_line(b)
        if not c or not cur_proc:
            continue
        t1 = clean(row[3]) if len(row) > 3 else None
        t2 = clean(row[5]) if len(row) > 5 else None
        t3 = clean(row[7]) if len(row) > 7 else None
        t4 = clean(row[9]) if len(row) > 9 else None
        def dim(idx):
            v = clean(row[idx]) if len(row) > idx else None
            if not v:
                return None
            m = re.search(r'(\d+(?:\.\d+)?)\s*mm', v)
            if m:
                return round(float(m.group(1)) / 10.0, 1)   # mm -> cm
            m = re.search(r'(\d+(?:\.\d+)?)', v)
            return round(float(m.group(1)), 1) if m else None
        ln, wd, dp = dim(11), dim(12), dim(13)
        upcoming = bool(re.search(r'upcoming|future', (cur_proc + ' ' + (c or '')), re.I))

        if cur_proc not in processes:
            processes[cur_proc] = {
                'process_name': cur_proc, 'process_group': group_for(cur_proc),
                'bay': None, 'max_len_cm': ln, 'max_wid_cm': wd, 'max_dep_cm': dp,
                'is_upcoming': upcoming,
            }
        caps.append({
            'capability_id': next_cap,
            'capability_ref': f'PCM-2026-{next_cap:04d}',
            'process_name': cur_proc, 'process_group': group_for(cur_proc),
            'customer_category': cur_cust or 'General', 'specification': c,
            'tier1_class': t1, 'tier2_class': t2, 'tier3_class': t3, 'tier4_class': t4,
            'max_len_cm': ln, 'max_wid_cm': wd, 'max_dep_cm': dp,
            'is_upcoming': upcoming,
        })
        next_cap += 1

    proc_list = sorted(processes.values(), key=lambda p: p['process_name'])
    customers = sorted({c['customer_category'] for c in caps})

    # revision history
    revs = []
    try:
        hs = wb['HIST']
        for row in hs.iter_rows(values_only=True):
            cells = [clean(x) for x in row]
            if cells and cells[1] and re.fullmatch(r'[A-Z]', str(cells[1] or '')):
                revs.append({'document_no': 'ATC-PCM-001', 'revision': cells[1],
                             'rev_date': str(cells[0]) if cells[0] else None,
                             'originator': cells[2], 'change_details': cells[3], 'reason': None})
    except Exception:
        pass
    if not revs:
        revs = [{'document_no': 'ATC-PCM-001', 'revision': 'A', 'rev_date': '2026-01-07',
                 'originator': 'JF TEH', 'change_details': 'Initial release', 'reason': '-'}]

    summary = {
        'total_capabilities': len(caps),
        'processes': len(proc_list),
        'customers': len(customers),
        'upcoming_services': sum(1 for c in caps if c['is_upcoming']),
        'total': len(caps),
    }

    demo = {
        '/mod28/alerts/summary': summary,
        '/mod28/processes': {'items': proc_list, 'total': len(proc_list)},
        '/mod28/capabilities': {'items': caps, 'total': len(caps)},
        '/mod28/revisions': {'items': revs, 'total': len(revs)},
    }
    out = os.path.join(os.path.dirname(__file__), 'pcm_demo.json')
    json.dump(demo, open(out, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
    print(f'capabilities={len(caps)} processes={len(proc_list)} customers={len(customers)} upcoming={summary["upcoming_services"]}')
    print('wrote', out)

if __name__ == '__main__':
    main()
